import pandas as pd
import openpyxl
import excel
import datetime as dt

# <1> Open workbook to read cell values
book = openpyxl.load_workbook("sales_data/stores.xlsx", data_only=True)


# <2> Get a worksheet object by name or index
sheet = book["2019"]
sheet = book.worksheets[0]

# <3> Get a list with all sheet names 

print(book.sheetnames)


# <4> Loop through the sheet objects.

for i in book.worksheets:
    print(i.title)

# <5> Getting the dimensions of the sheet

print(sheet.max_row) 
print(sheet.max_column)


# <6> Read the value of a single cell

print(sheet["B6"].value)
print(sheet.cell(row=6, column=2).value)


# <7> Read a range of values by using the excel module 

data = excel.read(book["2019"], (2,2), (8,6))
print(data[:2])


