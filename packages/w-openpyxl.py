import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, colors
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fills import PatternFill
import datetime as dt
import excel

#<1> instatiate a workbook

book = openpyxl.Workbook()


# <2> Get the first sheet and give it a name

sheet = book.active
sheet.title = "Sheet1"

# <3> Writting individual cells

sheet["A1"].value = "Hello 1"
print(sheet.cell(row=2, column=1, value="Hello 2"))


# <4> Formatting: fill color, alignment, border and font 


font_format = Font(color="FF0000", bold=True)
thin = Side(border_style="thin", color="FF0000")
sheet["A3"].value = "Hello 2"
sheet["A3"].font = font_format
sheet["A3"].border = Border(top=thin, right=thin, left=thin, bottom=thin)
sheet["A3"].alignment = Alignment(horizontal="center")
sheet["A3"].fill = PatternFill(fgColor = "FF0000", fill_type="solid")


# <5> Number formatting 

sheet["A4"].value = "4"
sheet["A4"].number_format = 0 

# <6> Date formatting 

sheet["A5"].value = dt.date(2016, 10, 13)
sheet["A5"].number_format = "mm/dd/yy"


# <7> Add image 

#sheet.add_image("images/python.png", "C1")

# <8> Two-dimensional sheet

data = [[None, "North", "South"], ["Last year", 3, 6], ["This year", 3, 6]]
excel.write(sheet,data, "A10")

# <9> Chart

chart = BarChart()
chart.type = "col"
chart.title = "Sales per region"
chart.x_axis.title = "Regions"
chart_data = Reference(sheet, min_row=11, min_col=1, max_row=12, max_col=3)
chart_categories = Reference(sheet, min_row=10, min_col=2, max_row=10, max_col=3)

# <10> From_rows interprate the data in the same way 

chart.add_data(chart_data, titles_from_data=True, from_rows=True)
chart.set_categories(chart_categories)
sheet.add_chart(chart, "A15")

# Saving the workbook creates the file on disk 


book = openpyxl.Workbook()
sheet = book.active
book.save("openpyxl.xlsx")


