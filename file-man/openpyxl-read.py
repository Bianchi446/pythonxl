import pandas as pd
import openpyxl
import excel
import datetime as dt

book = openpyxl.load_workbook("xl/stores.xlsx", data_only=True)

# Get a woorksheet object by name or index

sheet = book["2019"]
sheet = book.worksheets[0]

# Get a list with all sheet names

print(book.sheetnames)

# Looping  trough the sheet objects


for i in book.worksheets:
    print(i.title)


# Getting the dimensions

print(sheet.max_row, sheet.max_column)


# Read the value of a single cell - two ways

print(sheet["B6"].value)

print(sheet.cell(row=6, column=2).value)

